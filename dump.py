import os
import re
import io
import base64
from datetime import datetime

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup  # остаётся на случай, если понадобится fallback
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ——— Встроенные учётные данные ———
WORK_LOGIN = "ametrinhr@gmail.com"
WORK_PASSWORD = "95#Ametrin1995"

ROBOTAUA_LOGIN = "ametrinhr@gmail.com"
ROBOTAUA_PASSWORD = "95#Ametrin1995"

# ——— Глобальная переменная для хранения JWT robota.ua ———
ROBOTA_UA_TOKEN = ""

# ——— Константы ———
WORK_LOCALE = "uk_UA"
USER_AGENT = "StreamlitApp (ametrinhr@gmail.com)"

# Имя файла на диске, в котором будет храниться вся история
HISTORY_FILE = "resumes_history.xlsx"


# =============================================================================
# 1) Функции для авторизации и вытаскивания данных из API robota.ua
# =============================================================================

def robota_ua_login(username: str, password: str) -> str:
    """
    Авторизуемся на robota.ua → получаем Bearer-токен (JWT).
    POST https://auth-api.robota.ua/Login
    """
    url = "https://auth-api.robota.ua/Login"
    payload = {
        "username": username,
        "password": password
    }
    headers = {
        "User-Agent": USER_AGENT
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        raw = resp.json()
        if isinstance(raw, str):
            return raw.strip('"')
        if "accessToken" in raw:
            return raw["accessToken"]
        return ""
    except Exception as e:
        st.error(f"Не удалось получить JWT от robota.ua: {e}")
        return ""


def ensure_robota_ua_token() -> str:
    """
    Проверяем, есть ли глобальный ROBOTA_UA_TOKEN. Если нет — логинимся.
    Возвращаем JWT или пустую строку.
    """
    global ROBOTA_UA_TOKEN
    if not ROBOTA_UA_TOKEN:
        ROBOTA_UA_TOKEN = robota_ua_login(ROBOTAUA_LOGIN, ROBOTAUA_PASSWORD)
        if not ROBOTA_UA_TOKEN:
            st.error("Не удалось получить токен robota.ua.")
    return ROBOTA_UA_TOKEN


def build_fio(data: dict) -> str:
    """
    Склеиваем ФИО: сначала 'surname', затем 'name', затем 'fatherName' (если есть).
    """
    parts = []
    if data.get("surname"):
        parts.append(data["surname"].strip())
    if data.get("name"):
        parts.append(data["name"].strip())
    if data.get("fatherName"):
        parts.append(data["fatherName"].strip())
    return " ".join(parts)


def view_applicant_detail(apply_id: int, resume_type: int) -> (str, str):
    """
    POST /apply/view/{apply_id}?resumeType={resume_type}
    Возвращаем (fio, phone) для конкретного отклика (apply_id).
    """
    token = ensure_robota_ua_token()
    if not token:
        return "", ""

    url = f"https://employer-api.robota.ua/apply/view/{apply_id}?resumeType={resume_type}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.post(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except Exception:
        return "", ""

    try:
        data = resp.json()
    except Exception:
        return "", ""

    fio = build_fio(data)
    phone_raw = data.get("phone", "").strip()
    if not phone_raw:
        for c in data.get("contacts", []):
            if c.get("typeId") == "Phone":
                phone_raw = c.get("description", "").strip()
                break

    return fio, phone_raw


def get_resume_by_id_on_robotaua(resume_id: int, mark_view: bool = False) -> (str, str):
    """
    GET /resume/{resumeId}?markView={true/false}
    Возвращаем (fio, phone) для резюме с resume_id.
    """
    token = ensure_robota_ua_token()
    if not token:
        return "", ""

    url = f"https://employer-api.robota.ua/resume/{resume_id}"
    params = {"markView": str(mark_view).lower()}
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=15)
        resp.raise_for_status()
    except Exception:
        return "", ""

    try:
        data = resp.json()
    except Exception:
        return "", ""

    fio = build_fio(data)
    phone_raw = data.get("phone", "").strip()
    if not phone_raw:
        for c in data.get("contacts", []):
            if c.get("typeId") == "Phone":
                phone_raw = c.get("description", "").strip()
                break

    return fio, phone_raw


def extract_robotaua_candidate_id(url: str) -> str:
    """
    Ищем '/candidates/<digits>' в URL → возвращаем найденное число как строку.
    """
    m = re.search(r"/candidates/(\d+)", url)
    return m.group(1) if m else ""


def extract_robotaua_resume_id(url: str) -> str:
    """
    Ищем '/resume/<digits>' в URL → возвращаем найденное число как строку.
    """
    m = re.search(r"/resume/(\d+)", url)
    return m.group(1) if m else ""


def extract_applies_id(url: str) -> str:
    """
    Ищем параметр '?id=...' ← возвращаем всё после 'id=' до '&' или конца.
    """
    m = re.search(r"id=([^&]+)", url)
    return m.group(1) if m else ""


def extract_interaction_id(url: str) -> str:
    """
    Ищем '/interaction?id=...' ← возвращаем всё после 'id='.
    """
    m = re.search(r"/interaction\?id=([^&]+)", url)
    return m.group(1) if m else ""


def get_interaction_details(interaction_id: str) -> dict:
    """
    GET /apply/interaction/{interaction_id}
    Возвращает JSON с полями 'applyId', 'resumeId' и т.д.
    """
    token = ensure_robota_ua_token()
    if not token:
        return {}

    url = f"https://employer-api.robota.ua/apply/interaction/{interaction_id}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return {}


def parse_robota_ua_link(link: str) -> (str, str):
    """
    Универсальная функция: по ссылке на robota.ua (кандидат, резюме, отклик, interaction)
    возвращает (fio, телефон). Если ничего не удалось — ("", "").
    """
    # 1) /candidates/<id>
    cand_id = extract_robotaua_candidate_id(link)
    if cand_id:
        return get_resume_by_id_on_robotaua(int(cand_id), mark_view=True)

    # 2) /resume/<id>
    resume_id = extract_robotaua_resume_id(link)
    if resume_id:
        return get_resume_by_id_on_robotaua(int(resume_id), mark_view=True)

    # 3) /my/vacancies/.../applies?id=1234-some
    applies_str = extract_applies_id(link)
    if applies_str:
        m = re.match(r"(\d+)", applies_str)
        if m:
            apply_id = int(m.group(1))
            return view_applicant_detail(apply_id, resume_type=2)

    # 4) /apply/interaction/{interaction_id}
    interaction_id = extract_interaction_id(link)
    if interaction_id:
        details = get_interaction_details(interaction_id)
        apply_id = details.get("applyId", 0)
        resume_id2 = details.get("resumeId", 0)
        if apply_id:
            return view_applicant_detail(apply_id, resume_type=2)
        if resume_id2:
            return get_resume_by_id_on_robotaua(int(resume_id2), mark_view=True)

    # Нет распознанного формата — вернём пустые строки
    return "", ""


# =============================================================================
# 2) Функция сохранения/дозаписи истории в Excel
# =============================================================================

def save_history_to_excel(df: pd.DataFrame, path: str) -> None:
    """
    Сохраняет DataFrame в Excel-файл по указанному пути (path).
    Если файл уже существует, дозаписывает новые строки в конец.
    При этом: ширина колонок автоматически подогнана, шрифт Arial 11.
    """
    # Если файл существует — читаем предыдущую историю
    if os.path.exists(path):
        try:
            df_hist = pd.read_excel(path)
        except Exception:
            df_hist = pd.DataFrame()
        # Объединяем предыдущую историю и новые данные
        df_combined = pd.concat([df_hist, df], ignore_index=True)
    else:
        df_combined = df.copy()

    # Записываем объединённый DataFrame назад в Excel с форматированием
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_combined.to_excel(writer, index=False, sheet_name="Резюме")
        worksheet = writer.sheets["Резюме"]

        # Настраиваем ширину каждого столбца
        for col_idx, col_name in enumerate(df_combined.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                df_combined[col_name].astype(str).map(len).max(),
                len(col_name)
            )
            worksheet.column_dimensions[col_letter].width = max_length + 2

        # Устанавливаем шрифт Arial 11 для всех ячеек
        arial_11 = Font(name="Arial", size=11)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column
        ):
            for cell in row:
                cell.font = arial_11


# =============================================================================
# 3) Основной блок Streamlit-приложения
# =============================================================================

st.title("Сбор резюме с robota.ua и work.ua с историей")

links_input = st.text_area(
    "Вставьте ссылки на резюме (по одной на строку):",
    placeholder=(
        "https://robota.ua/candidates/22797009\n"
        "https://www.work.ua/resumes/9380298/"
    )
)

if st.button("Обработать"):
    links = [u.strip() for u in links_input.splitlines() if u.strip()]
    if not links:
        st.error("Нужно хотя бы одно URL.")
        st.stop()

    # Функция для форматирования телефона в единый вид: "0XX XXX XX XX"
    def format_phone(raw: str) -> str:
        digits = re.sub(r"\D", "", raw)
        if digits.startswith("380") and len(digits) == 12:
            digits = "0" + digits[3:]
        if len(digits) != 10:
            return raw
        return f"{digits[0:3]} {digits[3:6]} {digits[6:8]} {digits[8:10]}"

    # ——— Basic-авторизация для work.ua ———
    creds = f"{WORK_LOGIN}:{WORK_PASSWORD}"
    work_basic = base64.b64encode(creds.encode("utf-8")).decode("ascii")
    work_headers = {
        "Authorization": f"Basic {work_basic}",
        "X-Locale": WORK_LOCALE,
        "User-Agent": USER_AGENT
    }

    # ——— Получаем токен robota.ua один раз ———
    ensure_robota_ua_token()

    new_results = []
    for url in links:
        try:
            if "work.ua" in url.lower():
                # Работа через API Work.ua
                m = re.search(r"/resumes/(\d+)", url)
                if not m:
                    raise ValueError("Неверный формат URL work.ua")
                rid = m.group(1)

                r = requests.get(
                    "https://api.work.ua/resume",
                    params={"resume_id": rid},
                    headers=work_headers,
                    timeout=10
                )
                r.raise_for_status()
                data = r.json().get("result", {})
                fio = " ".join(filter(None, [data.get("first_name"), data.get("last_name")]))
                phone_raw = data.get("contacts", {}).get("phone_prim", "")

            else:
                # Работа через API Robota.ua
                fio, phone_raw = parse_robota_ua_link(url)
                if not fio:
                    st.warning(f"Не удалось получить ФИО по ссылке {url}, пропускаем")
                    continue

            phone = format_phone(phone_raw)
            date_str = datetime.now().strftime("%d.%m.%y")

            new_results.append({
                "Дата": date_str,
                "ФИО": fio,
                "Телефон": phone,
                "Ссылка": url
            })

        except Exception as e:
            st.warning(f"Ошибка обработки {url}: {e}")

    if new_results:
        df_new = pd.DataFrame(new_results)
        # Сохраняем (дописываем) в общий файл истории
        save_history_to_excel(df_new, HISTORY_FILE)

        # Выводим на экран всю историю целиком
        df_hist_all = pd.read_excel(HISTORY_FILE)
        st.subheader("Вся история кандидатов")
        st.table(df_hist_all)

        # Предоставляем пользователю возможность скачать файл-историю
        with open(HISTORY_FILE, "rb") as f:
            data_bytes = f.read()
        st.download_button(
            "Скачать историю (Excel)",
            data=data_bytes,
            file_name=f"resumes_history_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Нет успешно обработанных резюме.")
