import os
import re
import io
import base64
import logging
from datetime import datetime

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup  # остаётся на случай, если понадобится fallback
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ——— Отключаем «шумные» логи (ниже уровня ERROR) ———
logging.basicConfig(level=logging.ERROR)
logging.getLogger("urllib3").setLevel(logging.ERROR)
logging.getLogger("requests").setLevel(logging.ERROR)

# =============================================================================
# 0) Секреты: читаем логины/пароли из переменных окружения
# =============================================================================

WORK_LOGIN = "ametrinhr@gmail.com"
WORK_PASSWORD = "95#Ametrin1995"
ROBOTAUA_LOGIN = "ametrinhr@gmail.com"
ROBOTAUA_PASSWORD = "95#Ametrin1995"

if not WORK_LOGIN or not WORK_PASSWORD:
    st.error("Переменные окружения WORK_LOGIN и WORK_PASSWORD не заданы.")
    st.stop()

if not ROBOTAUA_LOGIN or not ROBOTAUA_PASSWORD:
    st.error("Переменные окружения ROBOTAUA_LOGIN и ROBOTAUA_PASSWORD не заданы.")
    st.stop()

# ——— Глобальная переменная для хранения JWT Robota.ua ———
ROBOTA_UA_TOKEN = ""

# ——— Константы ———
WORK_LOCALE = "uk_UA"
USER_AGENT = "StreamlitApp (ametrinhr@gmail.com)"
HISTORY_FILE = "resumes_history.xlsx"

# =============================================================================
# 1) Функции для авторизации и вытаскивания данных из API Robota.ua
# =============================================================================

def robota_ua_login(username: str, password: str) -> str:
    """
    Авторизуемся на Robota.ua → возвращаем JWT-токен (Bearer).
    """
    url = "https://auth-api.robota.ua/Login"
    payload = {"username": username, "password": password}
    headers = {"User-Agent": USER_AGENT}

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        raw = resp.json()
        # Иногда API возвращает строку, иногда — объект с ключом accessToken
        if isinstance(raw, str):
            return raw.strip('"')
        if "accessToken" in raw:
            return raw["accessToken"]
        return ""
    except Exception as e:
        st.error(f"Не удалось получить JWT от Robota.ua: {e}")
        return ""


def ensure_robota_ua_token() -> str:
    """
    Проверяем, есть ли глобальный ROBOTA_UA_TOKEN. Если нет — запускаем robota_ua_login.
    """
    global ROBOTA_UA_TOKEN
    if not ROBOTA_UA_TOKEN:
        ROBOTA_UA_TOKEN = robota_ua_login(ROBOTAUA_LOGIN, ROBOTAUA_PASSWORD)
        if not ROBOTA_UA_TOKEN:
            st.error("Не удалось получить токен Robota.ua.")
    return ROBOTA_UA_TOKEN


def build_fio(data: dict) -> str:
    """
    Склеиваем ФИО: «фамилия имя отчество» (если есть).
    """
    parts = []
    if data.get("surname"):
        parts.append(data["surname"].strip())
    if data.get("name"):
        parts.append(data["name"].strip())
    if data.get("fatherName"):
        parts.append(data["fatherName"].strip())
    return " ".join(parts)


def view_applicant_detail(apply_id: int, resume_type: int) -> (str, str):
    """
    Получаем детали отклика: POST /apply/view/{apply_id}?resumeType={resume_type}
    Возвращает (fio, phone).
    """
    token = ensure_robota_ua_token()
    if not token:
        return "", ""

    url = f"https://employer-api.robota.ua/apply/view/{apply_id}?resumeType={resume_type}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.post(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except Exception:
        return "", ""

    try:
        data = resp.json()
    except Exception:
        return "", ""

    fio = build_fio(data)
    phone_raw = data.get("phone", "").strip()
    if not phone_raw:
        for c in data.get("contacts", []):
            if c.get("typeId") == "Phone":
                phone_raw = c.get("description", "").strip()
                break

    return fio, phone_raw


def get_resume_by_id_on_robotaua(resume_id: int, mark_view: bool = False) -> (str, str):
    """
    Получаем резюме по ID: GET /resume/{resumeId}?markView={true/false}
    Возвращает (fio, phone).
    """
    token = ensure_robota_ua_token()
    if not token:
        return "", ""

    url = f"https://employer-api.robota.ua/resume/{resume_id}"
    params = {"markView": str(mark_view).lower()}
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=15)
        resp.raise_for_status()
    except Exception:
        return "", ""

    try:
        data = resp.json()
    except Exception:
        return "", ""

    fio = build_fio(data)
    phone_raw = data.get("phone", "").strip()
    if not phone_raw:
        for c in data.get("contacts", []):
            if c.get("typeId") == "Phone":
                phone_raw = c.get("description", "").strip()
                break

    return fio, phone_raw


def extract_robotaua_candidate_id(link: str) -> str:
    """
    Извлекает ID кандидата из URL "/candidates/<ID>".
    """
    m = re.search(r"/candidates/(\d+)", link)
    return m.group(1) if m else ""


def extract_robotaua_resume_id(link: str) -> str:
    """
    Извлекает ID резюме из URL "/resume/<ID>".
    """
    m = re.search(r"/resume/(\d+)", link)
    return m.group(1) if m else ""


def extract_applies_id(link: str) -> str:
    """
    Извлекает ID отклика из параметра "id=".
    """
    m = re.search(r"id=([^&]+)", link)
    return m.group(1) if m else ""


def extract_interaction_id(link: str) -> str:
    """
    Извлекает ID interaction из URL "/apply/interaction?id=<ID>".
    """
    m = re.search(r"/apply/interaction\?id=([^&]+)", link)
    return m.group(1) if m else ""


def get_interaction_details(interaction_id: str) -> dict:
    """
    Получаем детали interaction: GET /apply/interaction/{interaction_id}
    Возвращает JSON с полями 'applyId', 'resumeId' и т. д.
    """
    token = ensure_robota_ua_token()
    if not token:
        return {}

    url = f"https://employer-api.robota.ua/apply/interaction/{interaction_id}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": USER_AGENT
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return {}


def parse_robota_ua_link(link: str) -> (str, str):
    """
    Универсальный парсер ссылки на Robota.ua:
    1) /candidates/<ID>
    2) /resume/<ID>
    3) /my/vacancies/.../applies?id=...
    4) /apply/interaction?id=...
    Возвращает (fio, телефон) или ("", "").
    """
    # 1) /candidates/<id>
    cand_id = extract_robotaua_candidate_id(link)
    if cand_id:
        return get_resume_by_id_on_robotaua(int(cand_id), mark_view=True)

    # 2) /resume/<id>
    resume_id = extract_robotaua_resume_id(link)
    if resume_id:
        return get_resume_by_id_on_robotaua(int(resume_id), mark_view=True)

    # 3) /my/vacancies/.../applies?id=<число>[-....]
    applies_str = extract_applies_id(link)
    if applies_str:
        m = re.match(r"(\d+)", applies_str)
        if m:
            apply_id = int(m.group(1))
            return view_applicant_detail(apply_id, resume_type=2)

    # 4) /apply/interaction?id=<interaction_id>
    interaction_id = extract_interaction_id(link)
    if interaction_id:
        details = get_interaction_details(interaction_id)
        apply_id = details.get("applyId", 0)
        resume_id2 = details.get("resumeId", 0)
        if apply_id:
            return view_applicant_detail(apply_id, resume_type=2)
        if resume_id2:
            return get_resume_by_id_on_robotaua(int(resume_id2), mark_view=True)

    # Ничего не подошло
    return "", ""


# =============================================================================
# 2) Функция сохранения/дозаписи истории в Excel
# =============================================================================

def save_history_to_excel(df: pd.DataFrame, path: str) -> None:
    """
    Сохраняет DataFrame в указанный Excel-файл.
    Если файл существует, дозаписывает новые строки в конец.
    При записи автоматически подгоняет ширину колонок и задаёт шрифт Arial 11.
    """
    if os.path.exists(path):
        try:
            df_hist = pd.read_excel(path)
        except Exception:
            df_hist = pd.DataFrame()
        df_combined = pd.concat([df_hist, df], ignore_index=True)
    else:
        df_combined = df.copy()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_combined.to_excel(writer, index=False, sheet_name="Резюме")
        worksheet = writer.sheets["Резюме"]

        for col_idx, col_name in enumerate(df_combined.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                df_combined[col_name].astype(str).map(len).max(),
                len(col_name)
            )
            worksheet.column_dimensions[col_letter].width = max_length + 2

        arial_11 = Font(name="Arial", size=11)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column
        ):
            for cell in row:
                cell.font = arial_11


# =============================================================================
# 3) Основная часть Streamlit-приложения
# =============================================================================

st.title("Сбор резюме с robota.ua и work.ua с историей")

# Если ключа "processed" нет — инициализируем его как False
if "processed" not in st.session_state:
    st.session_state.processed = False

# Если ключа "show_confirm_clear" нет — инициализируем его как False
if "show_confirm_clear" not in st.session_state:
    st.session_state.show_confirm_clear = False

# Поле для ввода ссылок
links_input = st.text_area(
    "Вставьте ссылки на резюме (по одной на строку):",
    placeholder=(
        "https://robota.ua/candidates/22797009\n"
        "https://www.work.ua/resumes/9380298/"
    )
)

# Располагаем две колонки: "Обработать" и "Очистить файл"
col1, col2 = st.columns(2)

with col1:
    if st.button("Обработать"):
        links = [u.strip() for u in links_input.splitlines() if u.strip()]
        if not links:
            st.error("Нужно хотя бы одно URL.")
            st.stop()

        def format_phone(raw: str) -> str:
            """
            Форматируем телефон в вид "0XX XXX XX XX".
            Если длина цифр некорректна, возвращаем исходную строку.
            """
            digits = re.sub(r"\D", "", raw)
            if digits.startswith("380") and len(digits) == 12:
                digits = "0" + digits[3:]
            if len(digits) != 10:
                return raw
            return f"{digits[0:3]} {digits[3:6]} {digits[6:8]} {digits[8:10]}"

        # Basic-авторизация для Work.ua: формируем хедер один раз
        creds = f"{WORK_LOGIN}:{WORK_PASSWORD}"
        work_basic = base64.b64encode(creds.encode("utf-8")).decode("ascii")
        work_headers = {
            "Authorization": f"Basic {work_basic}",
            "X-Locale": WORK_LOCALE,
            "User-Agent": USER_AGENT
        }

        # Получаем JWT Robota.ua один раз
        ensure_robota_ua_token()

        new_results = []
        for url in links:
            try:
                if "work.ua" in url.lower():
                    # Обработка через API Work.ua
                    m = re.search(r"/resumes/(\d+)", url)
                    if not m:
                        raise ValueError("Неверный формат URL work.ua")
                    rid = m.group(1)

                    r = requests.get(
                        "https://api.work.ua/resume",
                        params={"resume_id": rid},
                        headers=work_headers,
                        timeout=10
                    )
                    r.raise_for_status()
                    data = r.json().get("result", {})
                    fio = " ".join(filter(None, [data.get("first_name"), data.get("last_name")]))
                    phone_raw = data.get("contacts", {}).get("phone_prim", "")

                else:
                    # Обработка через API Robota.ua
                    fio, phone_raw = parse_robota_ua_link(url)
                    if not fio:
                        st.warning(f"Не удалось получить ФИО по ссылке {url}, пропускаем")
                        continue

                phone = format_phone(phone_raw)
                date_str = datetime.now().strftime("%d.%m.%y")

                new_results.append({
                    "Дата": date_str,
                    "ФИО": fio,
                    "Телефон": phone,
                    "Ссылка": url
                })

            except Exception as e:
                st.warning(f"Ошибка обработки {url}: {e}")

        if new_results:
            # Сохраняем в Excel и ставим флаг processed = True
            df_new = pd.DataFrame(new_results)
            save_history_to_excel(df_new, HISTORY_FILE)
            st.session_state.processed = True
        else:
            st.info("Нет успешно обработанных резюме.")

with col2:
    if st.button("Очистить файл"):
        # Показываем модалку для подтверждения
        st.session_state.show_confirm_clear = True

# Блок подтверждения очистки (показывается поверх, без перезагрузки)
if st.session_state.show_confirm_clear:
    st.warning("Вы уверены, что хотите полностью очистить файл истории?")
    confirm1, confirm2 = st.columns(2)
    with confirm1:
        if st.button("Да, очистить", key="confirm_yes"):
            # Удаляем файл истории, если он существует
            if os.path.exists(HISTORY_FILE):
                try:
                    os.remove(HISTORY_FILE)
                    st.success("Файл истории успешно очищен.")
                except Exception as e:
                    st.error(f"Ошибка при удалении файла: {e}")
            else:
                st.info("Файл истории отсутствует, очищать нечего.")
            # Сбрасываем оба флага, чтобы модалка скрылась и таблица не отображалась
            st.session_state.show_confirm_clear = False
            st.session_state.processed = False

    with confirm2:
        if st.button("Нет", key="confirm_no"):
            # Пользователь отказался, просто скрываем модалку
            st.session_state.show_confirm_clear = False

# После всех кнопок и модалок: рендерим историю только если processed == True и файл существует
if st.session_state.processed and os.path.exists(HISTORY_FILE):
    try:
        df_hist_all = pd.read_excel(HISTORY_FILE)
        st.subheader("Вся история кандидатов")
        # Таблица занимает полную ширину, т.к. мы уже вышли из st.columns
        st.table(df_hist_all)

        # Кнопка скачивания истории
        with open(HISTORY_FILE, "rb") as f:
            data_bytes = f.read()
        st.download_button(
            "Скачать историю (Excel)",
            data=data_bytes,
            file_name=f"resumes_history_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Не удалось загрузить файл истории: {e}")
